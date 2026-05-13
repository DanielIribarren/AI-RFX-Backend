import os
from io import BytesIO

from openpyxl import load_workbook

os.environ.setdefault("SUPABASE_URL", "https://example.supabase.co")
os.environ.setdefault("SUPABASE_ANON_KEY", "test")
os.environ.setdefault("OPENAI_API_KEY", "sk-test1234567890")

from backend.models.apu_models import APUOutput
from backend.services.apu_generator import _APUExcelBuilder


def _sample_output() -> APUOutput:
    return APUOutput.model_validate({
        "rfx_id": "rfx-2024-0099",
        "project_name": "Urbanizacion Villas del Este",
        "client_company": "Inmobiliaria Central C.A.",
        "rfx_date": "2024-04-29",
        "tasa_bcv": 36.50,
        "tasa_bcv_missing": False,
        "pct_admin_gg": 0.16,
        "pct_utilidad": 0.10,
        "pct_sobre_costo_labor": 0.35,
        "partidas": [
            {
                "numero": "01",
                "descripcion": "Concreto f'c=210 kg/cm2 (3000 PSI) en columnas",
                "unidad": "m3",
                "cantidad_obra": 18.5,
                "rendimiento_und_dia": 6,
                "pct_sobre_costo_labor": 0.35,
                "materiales": [
                    {
                        "descripcion": "Cemento Portland 42.5 kg",
                        "unidad": "sa",
                        "cantidad": 7.0,
                        "desperdicio": 0.02,
                        "precio_unitario_usd": 8.50,
                        "es_precio_estimado": True,
                    },
                    {
                        "descripcion": "Arena lavada clasificada",
                        "unidad": "m3",
                        "cantidad": 0.45,
                        "desperdicio": 0.05,
                        "precio_unitario_usd": 28.00,
                        "es_precio_estimado": True,
                    },
                ],
                "equipos": [
                    {
                        "descripcion": "Concretera 1 saco",
                        "cantidad_dias": 1,
                        "costo_por_dia_usd": 45.00,
                        "dep_o_alq": 1.0,
                        "es_precio_estimado": True,
                    }
                ],
                "mano_obra": [
                    {
                        "descripcion": "Albanil oficial",
                        "cantidad_dias": 2,
                        "costo_por_dia_usd": 40.00,
                        "bono_usd": 8.00,
                        "es_precio_estimado": True,
                    },
                    {
                        "descripcion": "Ayudante de albanil",
                        "cantidad_dias": 2,
                        "costo_por_dia_usd": 20.00,
                        "bono_usd": 6.00,
                        "es_precio_estimado": True,
                    },
                ],
            }
        ],
        "warnings": ["Todos los precios son estimados con referencia de mercado venezolano."],
    })


def test_apu_chevron_builder_generates_expected_formulas():
    workbook_bytes = _APUExcelBuilder(_sample_output()).build()
    wb = load_workbook(BytesIO(workbook_bytes), data_only=False)

    assert wb.sheetnames == ["Resumen", "P01"]

    ws = wb["P01"]
    assert ws["G21"].value == "=+SUM(G12:G20)"
    assert ws["G22"].value == "=+G21"
    assert ws["G35"].value == "=(G34/$G$7)"
    assert ws["G48"].value == 0.35
    assert ws["G49"].value == "=+G47*G48"
    assert ws["G51"].value == "=+G50/$G$7"
    assert ws["G53"].value == "=+G51+G35+G22"
    assert ws["G54"].value == "=+G53*D54"
    assert ws["G56"].value == "=+G55*E56"
    assert ws["G58"].value == "=+G57"

    summary = wb["Resumen"]
    assert summary["E11"].value == "='P01'!$G$58"
    assert summary["F11"].value == "=D11*E11"
    assert summary["G11"].value == "=F11*$B$6"


def test_apu_chevron_builder_raises_when_section_exceeds_template_capacity():
    payload = _sample_output().model_dump()
    payload["partidas"][0]["materiales"] = [
        {
            "descripcion": f"Material {idx}",
            "unidad": "und",
            "cantidad": 1.0,
            "desperdicio": 0.0,
            "precio_unitario_usd": 1.0,
            "es_precio_estimado": True,
        }
        for idx in range(10)
    ]

    output = APUOutput.model_validate(payload)

    try:
        _APUExcelBuilder(output).build()
    except Exception as exc:
        assert "máximo Chevron" in str(exc)
    else:
        raise AssertionError("Expected Chevron capacity validation to fail")
