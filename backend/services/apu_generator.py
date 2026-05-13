"""
APU Generator Service — Construction Civil Venezuela.

Pipeline contract (deliberately stricter than proposal_generator.py):
  1. Load RFX from DB.
  2. Normalize products[] (handles the 4-naming-conventions mess in one place).
  3. Call LLM with JSON-mode + temperature 0.0.
  4. Validate output with Pydantic — fail loud on bad shape.
  5. One automatic retry feeding the validation error back to the model.
  6. Run business invariants (partidas non-empty, items per partida, etc.).
  7. Build Excel with openpyxl using Chevron geometry + native formulas.
  8. Upload to Supabase Storage and return a public URL.

Why this is NOT a copy of proposal_generator._call_ai:
  - proposal_generator returns broken HTML on retry failure (acceptable for visual).
  - We refuse to return a broken APU. Numerical errors propagate to budgets.
"""
from __future__ import annotations

import io
import json
import logging
import time
import uuid
from copy import copy
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from pydantic import ValidationError

from backend.core.config import get_openai_config
from backend.core.database import get_database_client
from backend.models.apu_models import (
    APUInput,
    APUItemCosto,
    APUItemEquipo,
    APUItemMO,
    APUItemMaterial,
    APUOutput,
    APUPartida,
    APUProductInput,
    APUResult,
)

logger = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────────────────────
# Constants
# ─────────────────────────────────────────────────────────────────────────────

PROMPT_VERSION = "1.3"
PROMPT_FILENAME = "apu_construction_ve.md"

LLM_TEMPERATURE = 0.0
LLM_TOP_P = 1.0
LLM_MAX_TOKENS = 16000  # gpt-4o ceiling; 30 partidas with full breakdowns can run ~14k chars
LLM_MAX_ATTEMPTS = 2  # first try + 1 retry
LLM_MAX_PRODUCTS_PER_BATCH = 6  # keep Chevron-sized outputs comfortably below truncation

STORAGE_BUCKET = "apu-exports"


# Excel style palette (mirrors the branding section in apu_construction_ve.md)
COLOR_HEADER_BG = "1F4E79"        # Azul acero
COLOR_TITLE_BG = "003366"          # Azul oscuro
COLOR_SECTION_BG = "D9E1F2"        # Azul grisáceo
COLOR_EDITABLE_BG = "FFF2CC"       # Amarillo (celda editable)
COLOR_TOTAL_BG = "FFE699"          # Amarillo más oscuro (total destacado)
COLOR_WHITE_TEXT = "FFFFFF"

FONT_NAME = "Arial"


class APUGenerationError(RuntimeError):
    """Raised when the APU pipeline cannot produce a valid output."""


# ─────────────────────────────────────────────────────────────────────────────
# Service
# ─────────────────────────────────────────────────────────────────────────────


class APUGeneratorService:
    """Generates an APU Excel file for a given RFX."""

    def __init__(self) -> None:
        self.openai_config = get_openai_config()
        self.db_client = get_database_client()
        self.openai_client = None
        self._prompt_text = self._load_prompt()

    # -- public ---------------------------------------------------------------

    def generate(
        self,
        rfx_id: str,
        tasa_bcv: Optional[float] = None,
        pct_admin_gg: Optional[float] = None,
        pct_costos_indirectos: Optional[float] = None,
        pct_utilidad: Optional[float] = None,
        pct_sobre_costo_labor: Optional[float] = None,
    ) -> APUResult:
        """Run the full pipeline. Raises APUGenerationError on failure.

        Optional overrides take precedence over rfx_data values and defaults:
            tasa_bcv  → BCV rate to use (None = use 0 and flag as missing)
            pct_admin_gg / pct_costos_indirectos → fraction in [0,1]
            pct_utilidad → fraction in [0,1] (None = 0.10 default)
            pct_sobre_costo_labor → labor social factor as fraction over base
                                     cost (e.g. 6.6091 = 660.91%)
        """
        started_at = time.time()
        logger.info("APU generation started for rfx_id=%s", rfx_id)

        rfx_data = self._load_rfx(rfx_id)
        apu_input = self._normalize_to_apu_input(
            rfx_id,
            rfx_data,
            tasa_bcv=tasa_bcv,
            pct_admin_gg=pct_admin_gg,
            pct_costos_indirectos=pct_costos_indirectos,
            pct_utilidad=pct_utilidad,
            pct_sobre_costo_labor=pct_sobre_costo_labor,
        )

        if not apu_input.products:
            raise APUGenerationError(
                f"RFX {rfx_id} has no products. Cannot generate APU."
            )

        output, attempts = self._call_llm_batched_with_validation(apu_input)

        excel_bytes = _APUExcelBuilder(output).build()
        excel_url, storage_path = self._persist(rfx_id, excel_bytes)
        self._update_rfx_apu_metadata(
            rfx_id=rfx_id,
            excel_url=excel_url,
            storage_path=storage_path,
            partidas_count=len(output.partidas),
        )

        elapsed = time.time() - started_at
        logger.info(
            "APU generation completed rfx_id=%s attempts=%d partidas=%d elapsed=%.2fs",
            rfx_id, attempts, len(output.partidas), elapsed,
        )

        return APUResult(
            rfx_id=rfx_id,
            excel_url=excel_url,
            excel_storage_path=storage_path,
            prompt_version=PROMPT_VERSION,
            llm_attempts=attempts,
            warnings=output.warnings,
            partidas_count=len(output.partidas),
        )

    # -- internals: data ------------------------------------------------------

    def _load_rfx(self, rfx_id: str) -> Dict[str, Any]:
        rfx_data = self.db_client.get_rfx_by_id(rfx_id)
        if not rfx_data:
            raise APUGenerationError(f"RFX {rfx_id} not found")

        # Products live in a separate table (rfx_products) and are not embedded
        # by get_rfx_by_id. Mirror the proposal_generator pattern and merge them.
        products = self.db_client.get_rfx_products(rfx_id) or []
        rfx_data["productos"] = products
        return rfx_data

    def _normalize_to_apu_input(
        self,
        rfx_id: str,
        rfx_data: Dict[str, Any],
        tasa_bcv: Optional[float] = None,
        pct_admin_gg: Optional[float] = None,
        pct_costos_indirectos: Optional[float] = None,
        pct_utilidad: Optional[float] = None,
        pct_sobre_costo_labor: Optional[float] = None,
    ) -> APUInput:
        """Resolve naming variants into the canonical prompt input.

        The service still accepts the legacy `pct_costos_indirectos` name, but
        internally we normalize everything to the Chevron-aligned v1.3 field
        `pct_admin_gg`.
        """
        productos = (
            rfx_data.get("productos")
            or rfx_data.get("products")
            or []
        )

        normalized: List[APUProductInput] = []
        for idx, p in enumerate(productos):
            normalized.append(
                APUProductInput(
                    id=str(p.get("id") or p.get("product_id") or f"p{idx + 1}"),
                    name=str(
                        p.get("product_name")
                        or p.get("name")
                        or p.get("nombre")
                        or f"Partida {idx + 1}"
                    ),
                    description=p.get("description") or p.get("descripcion"),
                    quantity=_safe_float(
                        p.get("quantity") or p.get("cantidad")
                    ),
                    unit=(
                        p.get("unit_of_measure")
                        or p.get("unit")
                        or p.get("unidad")
                    ),
                    specifications=p.get("specifications") or p.get("especificaciones"),
                    notes=p.get("notes") or p.get("notas"),
                    estimated_unit_price=_safe_float(
                        p.get("estimated_unit_price")
                        or p.get("precio_unitario_estimado")
                    ),
                    unit_cost=_safe_float(
                        p.get("unit_cost")
                        or p.get("precio_unitario")
                        or p.get("precio")
                    ),
                )
            )

        pct_admin_gg_value = (
            pct_admin_gg
            if pct_admin_gg is not None
            else pct_costos_indirectos
        )
        if pct_admin_gg_value is None:
            pct_admin_gg_value = (
                _safe_float(rfx_data.get("pct_admin_gg"))
                or _safe_float(rfx_data.get("pct_costos_indirectos"))
                or 0.22
            )

        return APUInput(
            rfx_id=rfx_id,
            project_name=str(
                rfx_data.get("project_name")
                or rfx_data.get("title")
                or rfx_data.get("nombre_proyecto")
                or "Proyecto sin nombre"
            ),
            client_company=str(
                rfx_data.get("client_company")
                or rfx_data.get("company_name")
                or rfx_data.get("requester_company")
                or (rfx_data.get("companies") or {}).get("name")
                or "Cliente sin nombre"
            ),
            rfx_date=str(
                rfx_data.get("rfx_date")
                or rfx_data.get("created_at")
                or datetime.now(timezone.utc).date().isoformat()
            )[:10],
            tasa_bcv=(
                tasa_bcv
                if tasa_bcv is not None
                else _safe_float(rfx_data.get("tasa_bcv")) or 0.0
            ),
            pct_admin_gg=pct_admin_gg_value,
            pct_utilidad=(
                pct_utilidad
                if pct_utilidad is not None
                else _safe_float(rfx_data.get("pct_utilidad")) or 0.10
            ),
            pct_sobre_costo_labor=(
                pct_sobre_costo_labor
                if pct_sobre_costo_labor is not None
                else _safe_float(rfx_data.get("pct_sobre_costo_labor")) or 6.6091
            ),
            products=normalized,
        )

    # -- internals: LLM -------------------------------------------------------

    def _call_llm_with_validation(self, apu_input: APUInput) -> Tuple[APUOutput, int]:
        """Strict pipeline: 1 try + 1 retry that includes the validation error.

        Three failure modes feed the retry loop:
          1. JSON parse error (output is not valid JSON)
          2. Pydantic schema error (shape/types/bounds wrong)
          3. Business invariant error (e.g., partida with all components empty)

        All three feed the same retry mechanism so the LLM can self-correct when
        we describe what went wrong.
        """
        last_error: Optional[str] = None
        last_raw_text: Optional[str] = None

        for attempt in range(1, LLM_MAX_ATTEMPTS + 1):
            raw_text = self._call_llm(
                apu_input,
                previous_attempt_text=last_raw_text,
                previous_error=last_error if attempt > 1 else None,
            )

            try:
                raw_dict = json.loads(raw_text)
            except json.JSONDecodeError as e:
                last_raw_text = raw_text
                last_error = f"Output is not valid JSON: {e.msg} at position {e.pos}"
                logger.warning(
                    "APU attempt %d: JSON parse failed (%s)", attempt, last_error
                )
                continue

            try:
                output = APUOutput.model_validate(raw_dict)
            except ValidationError as e:
                last_raw_text = raw_text
                last_error = self._summarize_validation_errors(e)
                logger.warning(
                    "APU attempt %d: schema validation failed (%s)", attempt, last_error
                )
                continue

            try:
                self._assert_invariants(output)
            except APUGenerationError as e:
                last_raw_text = raw_text
                last_error = f"Business invariant violated: {e}"
                logger.warning(
                    "APU attempt %d: invariants failed (%s)", attempt, last_error
                )
                continue

            return output, attempt

        raise APUGenerationError(
            f"LLM output failed validation after {LLM_MAX_ATTEMPTS} attempts. "
            f"Last error: {last_error}"
        )

    def _call_llm_batched_with_validation(
        self,
        apu_input: APUInput,
    ) -> Tuple[APUOutput, int]:
        """Generate APU output in product batches to avoid truncated JSON.

        The model remains the "brain" of the system, but we stop asking it to
        serialize 30 Chevron-style partidas in one response. Each batch is
        validated independently, then merged into one canonical APUOutput.
        """
        batches = self._chunk_apu_input(apu_input)
        if len(batches) == 1:
            return self._call_llm_with_validation(apu_input)

        logger.info(
            "APU generation split into %d LLM batches for rfx_id=%s (%d products)",
            len(batches),
            apu_input.rfx_id,
            len(apu_input.products),
        )

        merged_partidas: List[APUPartida] = []
        merged_warnings: List[str] = []
        total_attempts = 0
        first_output: Optional[APUOutput] = None

        for batch_idx, batch_input in enumerate(batches, start=1):
            logger.info(
                "APU batch %d/%d for rfx_id=%s (%d products)",
                batch_idx,
                len(batches),
                batch_input.rfx_id,
                len(batch_input.products),
            )
            batch_output, batch_attempts = self._call_llm_with_validation(batch_input)
            total_attempts += batch_attempts

            if first_output is None:
                first_output = batch_output

            merged_partidas.extend(batch_output.partidas)
            merged_warnings.extend(batch_output.warnings)

        if first_output is None:
            raise APUGenerationError("APU batching produced no outputs")

        merged_partidas = self._renumber_partidas(merged_partidas)
        merged_warnings = _dedupe_preserve_order(merged_warnings)

        merged_output = APUOutput(
            rfx_id=first_output.rfx_id,
            project_name=first_output.project_name,
            client_company=first_output.client_company,
            rfx_date=first_output.rfx_date,
            tasa_bcv=first_output.tasa_bcv,
            tasa_bcv_missing=first_output.tasa_bcv_missing,
            pct_admin_gg=first_output.pct_admin_gg,
            pct_utilidad=first_output.pct_utilidad,
            pct_sobre_costo_labor=first_output.pct_sobre_costo_labor,
            pct_iva=first_output.pct_iva,
            partidas=merged_partidas,
            warnings=merged_warnings,
        )
        self._assert_invariants(merged_output)
        return merged_output, total_attempts

    def _chunk_apu_input(self, apu_input: APUInput) -> List[APUInput]:
        if len(apu_input.products) <= LLM_MAX_PRODUCTS_PER_BATCH:
            return [apu_input]

        chunks: List[APUInput] = []
        for start in range(0, len(apu_input.products), LLM_MAX_PRODUCTS_PER_BATCH):
            chunk_products = apu_input.products[start:start + LLM_MAX_PRODUCTS_PER_BATCH]
            chunks.append(
                APUInput(
                    rfx_id=apu_input.rfx_id,
                    project_name=apu_input.project_name,
                    client_company=apu_input.client_company,
                    rfx_date=apu_input.rfx_date,
                    tasa_bcv=apu_input.tasa_bcv,
                    pct_admin_gg=apu_input.pct_admin_gg,
                    pct_utilidad=apu_input.pct_utilidad,
                    pct_sobre_costo_labor=apu_input.pct_sobre_costo_labor,
                    products=chunk_products,
                )
            )
        return chunks

    def _renumber_partidas(self, partidas: List[APUPartida]) -> List[APUPartida]:
        return [
            partida.model_copy(update={"numero": f"{idx:02d}"})
            for idx, partida in enumerate(partidas, start=1)
        ]

    def _call_llm(
        self,
        apu_input: APUInput,
        previous_attempt_text: Optional[str] = None,
        previous_error: Optional[str] = None,
    ) -> str:
        client = self._get_openai_client()

        messages: List[Dict[str, str]] = [
            {"role": "system", "content": self._prompt_text},
            {
                "role": "user",
                "content": apu_input.model_dump_json(indent=2),
            },
        ]

        if previous_attempt_text and previous_error:
            messages.append({"role": "assistant", "content": previous_attempt_text})
            messages.append({
                "role": "user",
                "content": (
                    "El JSON anterior falló validación. Errores:\n"
                    f"{previous_error}\n\n"
                    "Devuelve un JSON corregido que cumpla el contrato exactamente. "
                    "Solo el JSON, sin explicaciones, sin bloques de código."
                ),
            })

        response = client.chat.completions.create(
            model=self.openai_config.model,
            messages=messages,
            response_format={"type": "json_object"},
            temperature=LLM_TEMPERATURE,
            top_p=LLM_TOP_P,
            max_tokens=LLM_MAX_TOKENS,
        )

        choice = response.choices[0]
        content = (choice.message.content or "").strip()

        # When the model hits max_tokens it returns truncated (invalid) JSON.
        # Surface this loudly so the retry loop can attribute the failure correctly
        # rather than blaming the prompt.
        if choice.finish_reason == "length":
            logger.warning(
                "LLM output truncated (finish_reason=length, completion_tokens=%s). "
                "Increase LLM_MAX_TOKENS or reduce input partidas.",
                response.usage.completion_tokens if response.usage else "?",
            )

        return content

    def _get_openai_client(self):
        if self.openai_client is None:
            from openai import OpenAI
            self.openai_client = OpenAI(
                api_key=self.openai_config.api_key,
                max_retries=0,
            )
        return self.openai_client

    def _summarize_validation_errors(self, err: ValidationError) -> str:
        """Compact, model-friendly error summary."""
        lines = []
        for e in err.errors():
            loc = ".".join(str(x) for x in e["loc"])
            lines.append(f"- {loc}: {e['msg']}")
        return "\n".join(lines[:20])  # cap to keep retry payload bounded

    # -- internals: invariants ------------------------------------------------

    def _assert_invariants(self, output: APUOutput) -> None:
        if not output.partidas:
            raise APUGenerationError(
                "LLM returned zero partidas. Refusing to build empty APU."
            )

        for idx, partida in enumerate(output.partidas, start=1):
            total_items = (
                len(partida.materiales)
                + len(partida.mano_obra)
                + len(partida.equipos)
            )
            if total_items == 0:
                raise APUGenerationError(
                    f"Partida {idx} '{partida.descripcion}' has no items "
                    f"in any of materiales/mano_obra/equipos."
                )

    # -- internals: storage ---------------------------------------------------

    def _update_rfx_apu_metadata(
        self,
        rfx_id: str,
        excel_url: str,
        storage_path: str,
        partidas_count: int,
    ) -> None:
        """Track the latest APU on rfx_v2 so the UI can rehydrate without regenerating.

        Strict-mode: if this fails, we raise. Returning a successful APUResult while
        leaving the DB in an inconsistent state would mean the next page load
        wouldn't show the APU even though the file is in storage.
        """
        try:
            self.db_client.client.table("rfx_v2").update({
                "apu_excel_url": excel_url,
                "apu_excel_storage_path": storage_path,
                "apu_generated_at": datetime.now(timezone.utc).isoformat(),
                "apu_partidas_count": partidas_count,
                "apu_prompt_version": PROMPT_VERSION,
            }).eq("id", rfx_id).execute()
        except Exception as e:
            logger.error("APU metadata persist failed for rfx=%s: %s", rfx_id, e)
            raise APUGenerationError(
                f"APU generated but rfx_v2 metadata update failed: {e}"
            ) from e

    def _persist(self, rfx_id: str, excel_bytes: bytes) -> Tuple[str, str]:
        ts = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")
        storage_path = f"{rfx_id}/apu-{ts}-{uuid.uuid4().hex[:8]}.xlsx"

        public_url = self.db_client.upload_file_to_storage(
            bucket=STORAGE_BUCKET,
            file_path=storage_path,
            file_data=excel_bytes,
            content_type=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
        )
        return public_url, storage_path

    # -- internals: prompt loading -------------------------------------------

    def _load_prompt(self) -> str:
        prompt_path = (
            Path(__file__).resolve().parent.parent / "prompts" / PROMPT_FILENAME
        )
        if not prompt_path.exists():
            raise APUGenerationError(
                f"APU prompt file not found at {prompt_path}"
            )
        return prompt_path.read_text(encoding="utf-8")


# ─────────────────────────────────────────────────────────────────────────────
# Excel builder
# ─────────────────────────────────────────────────────────────────────────────


def _safe_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


# Chevron sheet geometry is fixed. Each partida gets its own sheet and the
# workbook starts with a simple Resumen sheet that centralizes BCV and totals
# without polluting the Chevron visual layout.
CHEVRON_MATERIAL_ROWS = range(12, 21)   # 9 slots
CHEVRON_EQUIPO_ROWS = range(25, 34)     # 9 slots
CHEVRON_MO_ROWS = range(38, 47)         # 9 slots
CHEVRON_MAX_ITEMS = 9
SUMMARY_SHEET_NAME = "Resumen"
DEFAULT_LOGO_PATH = (
    Path(__file__).resolve().parent.parent / "static" / "default" / "sabra_logo.png"
)
CHEVRON_TEMPLATE_PATH = (
    Path(__file__).resolve().parent.parent / "assets" / "apu_chevron_template.xlsx"
)

CHEVRON_COLUMN_WIDTHS = {
    "A": 12.36328125,
    "B": 36.90625,
    "C": 15.26953125,
    "D": 12.1796875,
    "E": 15.1796875,
    "F": 13.0,
    "G": 16.1796875,
    "H": 3.54296875,
    "I": 12.1796875,
    "J": 12.1796875,
    "K": 13.0,
    "L": 11.453125,
}

CHEVRON_ROW_HEIGHTS = {
    1: 42.5,
    4: 21.5,
    5: 26.0,
    6: 35.5,
    7: 14.5,
    8: 44.5,
    9: 7.0,
    10: 18.5,
    12: 14.5,
    13: 14.5,
    14: 14.5,
    15: 14.5,
    16: 14.5,
    17: 14.5,
    18: 14.5,
    19: 14.5,
    20: 14.5,
    21: 15.0,
    22: 14.5,
    23: 18.5,
    24: 14.5,
    25: 14.5,
    26: 14.5,
    27: 14.5,
    28: 14.5,
    29: 14.5,
    30: 14.5,
    31: 14.5,
    32: 14.5,
    33: 14.5,
    34: 14.5,
    35: 14.5,
    36: 18.5,
    37: 14.5,
    38: 14.5,
    39: 14.5,
    40: 14.5,
    41: 14.5,
    42: 14.5,
    43: 14.5,
    44: 14.5,
    45: 14.5,
    46: 14.5,
    50: 14.5,
    51: 14.5,
    58: 15.5,
    62: 15.5,
}


class _APUExcelBuilder:
    """Builds a Chevron-style workbook while preserving the current pipeline."""

    def __init__(self, output: APUOutput) -> None:
        self.output = output
        self.wb = Workbook()
        self.summary_ws = self.wb.active
        self.summary_ws.title = SUMMARY_SHEET_NAME
        self._logo_bytes = _load_default_logo_bytes()
        self._template_wb = self._load_chevron_template()
        self._template_ws = self._template_wb[self._template_wb.sheetnames[0]]

    def build(self) -> bytes:
        sheet_specs: List[Tuple[str, APUPartida]] = []
        for idx, partida in enumerate(self.output.partidas, start=1):
            ws = self.wb.create_sheet(title=self._sheet_name(idx))
            self._write_partida_sheet(ws, partida)
            sheet_specs.append((ws.title, partida))

        self._write_summary_sheet(sheet_specs)
        self._finalize_workbook_view()
        return self._to_bytes()

    def _load_chevron_template(self):
        if not CHEVRON_TEMPLATE_PATH.exists():
            raise APUGenerationError(
                f"Chevron APU template not found at {CHEVRON_TEMPLATE_PATH}"
            )
        return load_workbook(CHEVRON_TEMPLATE_PATH)

    def _write_partida_sheet(self, ws, partida: APUPartida) -> None:
        self._clone_template_sheet(ws)
        self._write_partida_header(ws, partida)
        self._write_materiales(ws, partida)
        self._write_equipos(ws, partida)
        self._write_mano_obra(ws, partida)
        self._write_partida_totals(ws, partida)

    def _clone_template_sheet(self, ws) -> None:
        src = self._template_ws

        for col, src_dim in src.column_dimensions.items():
            dst_dim = ws.column_dimensions[col]
            dst_dim.width = src_dim.width
            dst_dim.hidden = src_dim.hidden
            dst_dim.bestFit = src_dim.bestFit

        for row_idx, src_dim in src.row_dimensions.items():
            dst_dim = ws.row_dimensions[row_idx]
            dst_dim.height = src_dim.height
            dst_dim.hidden = src_dim.hidden

        for merge_ref in src.merged_cells.ranges:
            ws.merge_cells(str(merge_ref))

        max_row = src.max_row
        max_col = src.max_column
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                src_cell = src.cell(row=row, column=col)
                dst_cell = ws.cell(row=row, column=col)
                if src_cell.value is not None:
                    dst_cell.value = src_cell.value
                if src_cell.has_style:
                    dst_cell._style = copy(src_cell._style)
                if src_cell.number_format:
                    dst_cell.number_format = src_cell.number_format
                if src_cell.font:
                    dst_cell.font = copy(src_cell.font)
                if src_cell.fill:
                    dst_cell.fill = copy(src_cell.fill)
                if src_cell.border:
                    dst_cell.border = copy(src_cell.border)
                if src_cell.alignment:
                    dst_cell.alignment = copy(src_cell.alignment)
                if src_cell.protection:
                    dst_cell.protection = copy(src_cell.protection)

        ws.sheet_view.view = src.sheet_view.view
        ws.sheet_view.showGridLines = src.sheet_view.showGridLines
        ws.page_margins = copy(src.page_margins)
        ws.page_setup = copy(src.page_setup)
        ws.print_options = copy(src.print_options)
        ws.sheet_properties = copy(src.sheet_properties)
        ws.print_title_cols = src.print_title_cols
        ws.print_title_rows = src.print_title_rows
        ws.freeze_panes = src.freeze_panes
        if src.print_area:
            ws.print_area = src.print_area

        self._add_template_logo(ws)
        ws.sheet_view.view = "normal"

    def _add_template_logo(self, ws) -> None:
        if getattr(self._template_ws, "_images", None):
            for src_image in self._template_ws._images:
                try:
                    image_bytes = src_image._data()
                    image = XLImage(io.BytesIO(image_bytes))
                    image.width = src_image.width
                    image.height = src_image.height
                    anchor = src_image.anchor._from
                    coord = f"{chr(65 + anchor.col)}{anchor.row + 1}"
                    ws.add_image(image, coord)
                except Exception as e:
                    logger.warning(
                        "Could not clone template image into sheet %s: %s",
                        ws.title,
                        e,
                    )
        elif self._logo_bytes:
            try:
                image = XLImage(io.BytesIO(self._logo_bytes))
                image.width = 419
                image.height = 175
                ws.add_image(image, "A1")
            except Exception as e:
                logger.warning("Could not inject fallback logo into sheet %s: %s", ws.title, e)

    def _write_partida_header(self, ws, partida: APUPartida) -> None:
        self._set_text(ws, "F1", "Partida Nro.:", font_size=10, align="right")
        self._set_text(ws, "G1", partida.numero, bold=True, font_size=12, align="center")

        self._set_text(ws, "A3", "Obra:", font_size=11)
        self._set_text(ws, "B3", self.output.project_name, font_size=11)
        self._set_text(ws, "A4", "Contratante:", font_size=11)
        self._set_text(ws, "B4", self.output.client_company, font_size=11)

        self._set_text(
            ws,
            "A5",
            "ANALISIS DE PRECIOS UNITARIOS",
            bold=True,
            font_size=20,
            align="center",
        )

        self._set_text(ws, "E6", "Cantidad:", font_size=11, align="center")
        self._set_text(ws, "F6", "Unidad:", font_size=11, align="center")
        self._set_text(ws, "G6", "Rendimiento (Und/dia):", font_size=10, align="center")

        self._set_text(ws, "A7", "Descripción de la Partida:", font_size=10)
        description_cell = ws["A8"]
        description_cell.value = partida.descripcion
        description_cell.font = Font(name=FONT_NAME, size=11)
        description_cell.alignment = Alignment(
            horizontal="left",
            vertical="top",
            wrap_text=True,
        )

        qty_cell = ws["E7"]
        qty_cell.value = partida.cantidad_obra
        qty_cell.number_format = "#,##0.00"
        qty_cell.font = Font(name=FONT_NAME, size=11, bold=True)
        qty_cell.alignment = Alignment(horizontal="center")
        qty_cell.border = _thin_border()

        unit_cell = ws["F7"]
        unit_cell.value = partida.unidad
        unit_cell.font = Font(name=FONT_NAME, size=11, bold=True)
        unit_cell.alignment = Alignment(horizontal="center")
        unit_cell.border = _thin_border()

        rendimiento_cell = ws["G7"]
        rendimiento_cell.value = partida.rendimiento_und_dia
        rendimiento_cell.number_format = "#,##0.00000"
        rendimiento_cell.font = Font(name=FONT_NAME, size=11, bold=True)
        rendimiento_cell.alignment = Alignment(horizontal="center")
        rendimiento_cell.border = _thin_border()

    def _write_materiales(self, ws, partida: APUPartida) -> None:
        items = [self._coerce_material(item) for item in partida.materiales]
        self._assert_section_capacity(partida, "materiales", len(items))

        self._set_text(ws, "A10", "1.-MATERIALES", bold=True, font_size=11)
        for coord, value in {
            "A11": "Ítem",
            "B11": "Descripción",
            "C11": "Unidad",
            "D11": "Cantidad",
            "E11": "Desperd.",
            "F11": "Costo Unitario",
            "G11": "Costo Total $",
        }.items():
            self._set_header_cell(ws, coord, value)

        for index, row_num in enumerate(CHEVRON_MATERIAL_ROWS, start=1):
            ws[f"A{row_num}"] = index
            ws[f"A{row_num}"].alignment = Alignment(horizontal="center")

            item = items[index - 1] if index <= len(items) else None
            if item:
                ws.row_dimensions[row_num].hidden = False
                self._set_text(ws, f"B{row_num}", _truncate(item["descripcion"]))
                self._set_text(ws, f"C{row_num}", item["unidad"], align="center")
                self._set_number(ws, f"D{row_num}", item["cantidad"], "#,##0.0000")
                self._set_number(ws, f"E{row_num}", item["desperdicio"], "0.00%")
                self._set_number(
                    ws,
                    f"F{row_num}",
                    item["precio_unitario_usd"],
                    "#,##0.00",
                    editable=item["es_precio_estimado"],
                )
            else:
                ws.row_dimensions[row_num].hidden = True

            ws[f"G{row_num}"] = f"=+D{row_num}*F{row_num}*(1+E{row_num})"
            ws[f"G{row_num}"].number_format = "#,##0.00"
            ws[f"I{row_num}"] = f"=+G{row_num}*$E$7"
            ws[f"I{row_num}"].number_format = "#,##0.00"

            for col in "ABCDEFG":
                ws[f"{col}{row_num}"].border = _thin_border()

        self._set_text(ws, "F21", "Total Materiales:", bold=True, font_size=11, align="right")
        ws["G21"] = "=+SUM(G12:G20)"
        ws["G21"].number_format = "#,##0.00"
        ws["G21"].border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"))

        self._set_text(
            ws,
            "F22",
            "Costo Unitario de Materiales:",
            bold=True,
            font_size=11,
            align="right",
        )
        ws["G22"] = "=+G21"
        ws["G22"].number_format = "#,##0.00"
        ws["G22"].border = _medium_border()

    def _write_equipos(self, ws, partida: APUPartida) -> None:
        items = [self._coerce_equipo(item) for item in partida.equipos]
        self._assert_section_capacity(partida, "equipos", len(items))

        self._set_text(ws, "A23", "2.EQUIPOS", bold=True, font_size=11)
        for coord, value in {
            "A24": "Ítem",
            "B24": "Descripción",
            "D24": "Cantidad",
            "E24": "Costo por dia",
            "F24": "Dep. o Alq.",
            "G24": "Costo Total $",
        }.items():
            self._set_header_cell(ws, coord, value)

        for index, row_num in enumerate(CHEVRON_EQUIPO_ROWS, start=1):
            ws[f"A{row_num}"] = index
            ws[f"A{row_num}"].alignment = Alignment(horizontal="center")

            item = items[index - 1] if index <= len(items) else None
            if item:
                ws.row_dimensions[row_num].hidden = False
                self._set_text(ws, f"B{row_num}", _truncate(item["descripcion"]))
                self._set_number(ws, f"D{row_num}", item["cantidad_dias"], "#,##0.00")
                self._set_number(
                    ws,
                    f"E{row_num}",
                    item["costo_por_dia_usd"],
                    "#,##0.00",
                    editable=item["es_precio_estimado"],
                )
                self._set_number(ws, f"F{row_num}", item["dep_o_alq"], "0.00")
            else:
                ws.row_dimensions[row_num].hidden = True

            ws[f"G{row_num}"] = f"=+D{row_num}*E{row_num}*F{row_num}"
            ws[f"G{row_num}"].number_format = "#,##0.00"
            ws[f"I{row_num}"] = f"=+G{row_num}*$E$7/$G$7"
            ws[f"I{row_num}"].number_format = "#,##0.00"

            for col in "ABCDEFG":
                ws[f"{col}{row_num}"].border = _thin_border()

        self._set_text(ws, "F34", "Total Equipos:", bold=True, font_size=11, align="right")
        ws["G34"] = "=SUM(G25:G33)"
        ws["G34"].number_format = "#,##0.00"
        ws["G34"].border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"))

        self._set_text(
            ws,
            "F35",
            "Costo Unitario de Equipos:",
            bold=True,
            font_size=11,
            align="right",
        )
        ws["G35"] = "=(G34/$G$7)"
        ws["G35"].number_format = "#,##0.00"
        ws["G35"].border = _medium_border()

    def _write_mano_obra(self, ws, partida: APUPartida) -> None:
        items = [self._coerce_mano_obra(item) for item in partida.mano_obra]
        self._assert_section_capacity(partida, "mano_obra", len(items))

        self._set_text(ws, "A36", "3.-LABOR DIRECTA (MANO DE OBRA)", bold=True, font_size=11)
        for coord, value in {
            "A37": "Ítem",
            "B37": "Descripción",
            "D37": "Cantidad",
            "E37": "Costo por dia",
            "F37": "Bono",
            "G37": "Costo Total $",
        }.items():
            self._set_header_cell(ws, coord, value)

        for index, row_num in enumerate(CHEVRON_MO_ROWS, start=1):
            ws[f"A{row_num}"] = index
            ws[f"A{row_num}"].alignment = Alignment(horizontal="center")

            item = items[index - 1] if index <= len(items) else None
            if item:
                ws.row_dimensions[row_num].hidden = False
                self._set_text(ws, f"B{row_num}", _truncate(item["descripcion"]))
                self._set_number(ws, f"D{row_num}", item["cantidad_dias"], "#,##0.00")
                self._set_number(
                    ws,
                    f"E{row_num}",
                    item["costo_por_dia_usd"],
                    "#,##0.00",
                    editable=item["es_precio_estimado"],
                )
                self._set_number(ws, f"F{row_num}", item["bono_usd"], "#,##0.00")
            else:
                ws.row_dimensions[row_num].hidden = True

            ws[f"G{row_num}"] = f"=ROUND(+E{row_num}*D{row_num},2)"
            ws[f"G{row_num}"].number_format = "#,##0.00"

            for col in "ABCDEFG":
                ws[f"{col}{row_num}"].border = _thin_border()

        ws["D47"] = "=SUM(D38:D46)"
        ws["D47"].number_format = "0.00"
        ws["D47"].font = Font(name=FONT_NAME, size=11, bold=True)
        ws["D47"].alignment = Alignment(horizontal="center")
        self._set_text(ws, "E47", "Sub-total mano de obra:", font_size=11, align="right")
        ws["G47"] = "=+SUM(G38:G46)"
        ws["G47"].number_format = "#,##0.00"
        ws["G47"].font = Font(name=FONT_NAME, size=11, bold=True)
        ws["G47"].alignment = Alignment(horizontal="right")
        ws["G47"].border = _thin_border()

        self._set_text(ws, "F48", "Sub-total labor(%):", font_size=11, align="right")
        self._set_number(
            ws,
            "G48",
            partida.pct_sobre_costo_labor,
            "0.00%",
            editable=True,
            bold=True,
            align="center",
        )
        ws["G48"].border = _thin_border()

        self._set_text(ws, "E49", "Sobre costo de Labor:", font_size=11, align="right")
        ws["G49"] = "=+G47*G48"
        ws["G49"].number_format = "#,##0.00"
        ws["G49"].alignment = Alignment(horizontal="center")
        ws["G49"].border = _thin_border()

        self._set_text(ws, "F50", "Total Mano de Obra:", font_size=11, align="right")
        ws["G50"] = "=+G47+G49"
        ws["G50"].number_format = "#,##0.00"
        ws["G50"].border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"))

        self._set_text(
            ws,
            "F51",
            "Costo Unitario Mano de Obra:",
            bold=True,
            font_size=11,
            align="right",
        )
        ws["G51"] = "=+G50/$G$7"
        ws["G51"].number_format = "#,##0.00"
        ws["G51"].font = Font(name=FONT_NAME, size=11, bold=True)
        ws["G51"].border = _medium_border()

    def _write_partida_totals(self, ws, partida: APUPartida) -> None:
        self._set_text(ws, "A52", "HH Totales:", font_size=11)
        ws["B52"] = "=+B53*SUM(D38:D46)*8"
        ws["B52"].number_format = "0.00"

        self._set_text(ws, "A53", "Duración: ", font_size=11)
        ws["B53"] = "=+E7/G7"
        ws["B53"].number_format = "0.00"

        self._set_text(ws, "F53", "Costo Directo por Unidad:", bold=True, font_size=11, align="right")
        ws["G53"] = "=+G51+G35+G22"
        ws["G53"].number_format = "#,##0.00"
        ws["G53"].border = _thin_border()

        self._set_number(ws, "D54", self.output.pct_admin_gg, "0%", editable=True)
        self._set_text(ws, "F54", " Administración y Gastos Gener.:", font_size=11, align="right")
        ws["G54"] = "=+G53*D54"
        ws["G54"].number_format = "#,##0.00"
        ws["G54"].border = _thin_border()

        self._set_text(ws, "F55", "Sub-Total:", font_size=11, align="right")
        ws["G55"] = "=+G53+G54"
        ws["G55"].number_format = "#,##0.00"
        ws["G55"].border = _thin_border()

        self._set_number(ws, "E56", self.output.pct_utilidad, "0%", editable=True)
        self._set_text(ws, "F56", "Utilidad e Impr.:", font_size=11, align="right")
        ws["G56"] = "=+G55*E56"
        ws["G56"].number_format = "#,##0.00"
        ws["G56"].border = _thin_border()

        self._set_text(ws, "F57", "Sub-Total:", font_size=11, align="right")
        ws["G57"] = "=+G55+G56"
        ws["G57"].number_format = "#,##0.00"
        ws["G57"].border = _thin_border()

        self._set_text(ws, "F58", "PRECIO UNITARIO $.", bold=True, font_size=11, align="right")
        ws["G58"] = "=+G57"
        ws["G58"].number_format = "#,##0.00"
        ws["G58"].font = Font(name=FONT_NAME, size=12, bold=True)
        ws["G58"].alignment = Alignment(horizontal="center")
        ws["G58"].border = _thin_border()

    def _write_summary_sheet(self, sheet_specs: List[Tuple[str, APUPartida]]) -> None:
        ws = self.summary_ws
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 50
        ws.column_dimensions["C"].width = 10
        ws.column_dimensions["D"].width = 14
        ws.column_dimensions["E"].width = 16
        ws.column_dimensions["F"].width = 18
        ws.column_dimensions["G"].width = 18

        ws.merge_cells("A1:G1")
        self._set_text(
            ws,
            "A1",
            "RESUMEN APU CHEVRON",
            bold=True,
            font_size=16,
            align="center",
        )
        self._set_text(ws, "A3", "Proyecto:", bold=True)
        self._set_text(ws, "B3", self.output.project_name)
        self._set_text(ws, "A4", "Cliente:", bold=True)
        self._set_text(ws, "B4", self.output.client_company)
        self._set_text(ws, "D3", "RFX ID:", bold=True)
        self._set_text(ws, "E3", self.output.rfx_id)
        self._set_text(ws, "D4", "Fecha:", bold=True)
        self._set_text(ws, "E4", self.output.rfx_date)

        self._set_text(ws, "A6", "Tasa BCV:", bold=True)
        self._set_number(
            ws,
            "B6",
            self.output.tasa_bcv if not self.output.tasa_bcv_missing else 0,
            "#,##0.00",
            editable=True,
        )
        self._set_text(ws, "D6", "Admin & GG:", bold=True)
        self._set_number(ws, "E6", self.output.pct_admin_gg, "0.00%", editable=True)
        self._set_text(ws, "D7", "Utilidad:", bold=True)
        self._set_number(ws, "E7", self.output.pct_utilidad, "0.00%", editable=True)
        self._set_text(ws, "D8", "SC Labor base:", bold=True)
        self._set_number(ws, "E8", self.output.pct_sobre_costo_labor, "0.00%", editable=True)

        for coord, value in {
            "A10": "Partida",
            "B10": "Descripción",
            "C10": "Unidad",
            "D10": "Cantidad",
            "E10": "P.U. USD",
            "F10": "Total USD",
            "G10": "Total VES",
        }.items():
            self._set_header_cell(ws, coord, value)

        row = 11
        for sheet_name, partida in sheet_specs:
            self._set_text(ws, f"A{row}", partida.numero)
            self._set_text(ws, f"B{row}", _truncate(partida.descripcion, 80))
            self._set_text(ws, f"C{row}", partida.unidad, align="center")
            self._set_number(ws, f"D{row}", partida.cantidad_obra, "#,##0.00")
            ws[f"E{row}"] = f"='{sheet_name}'!$G$58"
            ws[f"E{row}"].number_format = "#,##0.00"
            ws[f"F{row}"] = f"=D{row}*E{row}"
            ws[f"F{row}"].number_format = "#,##0.00"
            ws[f"G{row}"] = f"=F{row}*$B$6"
            ws[f"G{row}"].number_format = "#,##0.00"
            for col in "ABCDEFG":
                ws[f"{col}{row}"].border = _thin_border()
            row += 1

        self._set_text(ws, f"E{row}", "TOTAL", bold=True, align="right")
        ws[f"F{row}"] = f"=SUM(F11:F{row - 1})"
        ws[f"F{row}"].number_format = "#,##0.00"
        ws[f"G{row}"] = f"=SUM(G11:G{row - 1})"
        ws[f"G{row}"].number_format = "#,##0.00"
        for col in "EFG":
            ws[f"{col}{row}"].font = Font(name=FONT_NAME, bold=True)
            ws[f"{col}{row}"].border = _medium_border()

        if self.output.warnings:
            row += 2
            self._set_text(ws, f"A{row}", "Advertencias:", bold=True, font_size=11)
            row += 1
            for warning in self.output.warnings:
                ws.merge_cells(f"A{row}:G{row}")
                self._set_text(ws, f"A{row}", f"• {warning}", font_size=10)
                row += 1

        if self.output.tasa_bcv_missing:
            row += 1
            ws.merge_cells(f"A{row}:G{row}")
            self._set_text(
                ws,
                f"A{row}",
                "⚠ Tasa BCV no provista. Actualice la celda B6 para convertir el resumen a VES.",
                font_size=10,
            )
        ws.sheet_view.view = "normal"

    def _finalize_workbook_view(self) -> None:
        if len(self.wb.worksheets) > 1:
            self.wb.move_sheet(self.summary_ws, offset=len(self.wb.worksheets) - 1 - self.wb.index(self.summary_ws))
            self.wb.active = 0
        else:
            self.wb.active = self.wb.index(self.summary_ws)

    def _sheet_name(self, idx: int) -> str:
        return f"P{idx:02d}"

    def _assert_section_capacity(
        self,
        partida: APUPartida,
        section_name: str,
        item_count: int,
    ) -> None:
        if item_count > CHEVRON_MAX_ITEMS:
            raise APUGenerationError(
                f"Partida '{partida.descripcion}' excede el máximo Chevron de "
                f"{CHEVRON_MAX_ITEMS} filas en {section_name}: {item_count} ítems."
            )

    def _coerce_material(self, item: APUItemMaterial | APUItemCosto) -> Dict[str, Any]:
        if isinstance(item, APUItemMaterial):
            return {
                "descripcion": item.descripcion,
                "unidad": item.unidad,
                "cantidad": item.cantidad,
                "desperdicio": item.desperdicio,
                "precio_unitario_usd": item.precio_unitario_usd,
                "es_precio_estimado": item.es_precio_estimado,
            }
        return {
            "descripcion": item.descripcion,
            "unidad": item.unidad,
            "cantidad": item.cantidad,
            "desperdicio": 0.0,
            "precio_unitario_usd": item.precio_unitario_usd,
            "es_precio_estimado": item.es_precio_estimado,
        }

    def _coerce_equipo(self, item: APUItemEquipo | APUItemCosto) -> Dict[str, Any]:
        if isinstance(item, APUItemEquipo):
            return {
                "descripcion": item.descripcion,
                "cantidad_dias": item.cantidad_dias,
                "costo_por_dia_usd": item.costo_por_dia_usd,
                "dep_o_alq": item.dep_o_alq,
                "es_precio_estimado": item.es_precio_estimado,
            }
        return {
            "descripcion": item.descripcion,
            "cantidad_dias": item.cantidad,
            "costo_por_dia_usd": item.precio_unitario_usd,
            "dep_o_alq": 1.0,
            "es_precio_estimado": item.es_precio_estimado,
        }

    def _coerce_mano_obra(self, item: APUItemMO | APUItemCosto) -> Dict[str, Any]:
        if isinstance(item, APUItemMO):
            return {
                "descripcion": item.descripcion,
                "cantidad_dias": item.cantidad_dias,
                "costo_por_dia_usd": item.costo_por_dia_usd,
                "bono_usd": item.bono_usd,
                "es_precio_estimado": item.es_precio_estimado,
            }
        return {
            "descripcion": item.descripcion,
            "cantidad_dias": item.cantidad,
            "costo_por_dia_usd": item.precio_unitario_usd,
            "bono_usd": 0.0,
            "es_precio_estimado": item.es_precio_estimado,
        }

    def _set_header_cell(self, ws, coord: str, value: str) -> None:
        cell = ws[coord]
        cell.value = value
        cell.font = Font(name=FONT_NAME, size=11, bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.border = _thin_border()

    def _set_text(
        self,
        ws,
        coord: str,
        value: str,
        *,
        bold: bool = False,
        font_size: int = 11,
        align: Optional[str] = None,
    ) -> None:
        cell = ws[coord]
        cell.value = value
        cell.font = Font(name=FONT_NAME, size=font_size, bold=bold)
        if align:
            cell.alignment = Alignment(horizontal=align)

    def _set_number(
        self,
        ws,
        coord: str,
        value: float,
        fmt: str,
        *,
        editable: bool = False,
        bold: bool = False,
        align: Optional[str] = None,
    ) -> None:
        cell = ws[coord]
        cell.value = value
        cell.number_format = fmt
        cell.font = Font(name=FONT_NAME, size=11, bold=bold)
        if align:
            cell.alignment = Alignment(horizontal=align)
        if editable:
            cell.fill = PatternFill("solid", fgColor=COLOR_EDITABLE_BG)

    def _to_bytes(self) -> bytes:
        buf = io.BytesIO()
        self.wb.save(buf)
        return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────


def _thin_border() -> Border:
    side = Side(style="thin", color="CCCCCC")
    return Border(left=side, right=side, top=side, bottom=side)


def _medium_border() -> Border:
    side = Side(style="medium", color="000000")
    return Border(left=side, right=side, top=side, bottom=side)


def _load_default_logo_bytes() -> Optional[bytes]:
    if not DEFAULT_LOGO_PATH.exists():
        logger.warning("APU logo file not found at %s", DEFAULT_LOGO_PATH)
        return None
    try:
        return DEFAULT_LOGO_PATH.read_bytes()
    except Exception as e:
        logger.warning("Could not read APU logo file %s: %s", DEFAULT_LOGO_PATH, e)
        return None


def _truncate(value: str, max_length: int = 55) -> str:
    value = (value or "").strip()
    if len(value) <= max_length:
        return value
    return value[: max_length - 1].rstrip() + "…"


def _dedupe_preserve_order(values: List[str]) -> List[str]:
    seen = set()
    ordered: List[str] = []
    for value in values:
        if value in seen:
            continue
        seen.add(value)
        ordered.append(value)
    return ordered
